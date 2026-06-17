import glob
import re
from collections import Counter
from datetime import date, timedelta
from itertools import count

import mysql.connector as mc
import openpyxl

PLACEHOLDER_EMAILS = {'mobile', 'name@moh.gov.sa', 'nnnn@gmail.com'}
FEMALE_HINTS = ['ممرضة', 'أخصائية', 'سكرتيرة', 'منسقة', 'مشرفة', 'كاتبة', 'عاملة', 'فنية', 'محاسبة', 'مدخلة', 'قابلة']
MALE_HINTS = ['ممرض', 'أخصائي', 'سكرتير', 'منسق', 'مشرف', 'كاتب', 'عامل', 'فني', 'محاسب', 'مدير', 'سائق']


def locate_excel():
    xlsx_files = glob.glob('*.xlsx')
    if not xlsx_files:
        raise FileNotFoundError('No Excel file found in current directory')
    return xlsx_files[0]


def load_rows(excel_path):
    workbook = openpyxl.load_workbook(excel_path)
    sheet = workbook.active
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    rows = [row for row in sheet.iter_rows(min_row=2, values_only=True)]
    return headers, rows


def sanitize_string(value):
    if value is None:
        return ''
    return str(value).strip()


def sanitize_phone(value):
    phone = sanitize_string(value)
    if not phone:
        return '0000000000'
    digits = ''.join(ch for ch in phone if ch.isdigit() or ch == '+')
    if not digits:
        return '0000000000'
    return digits[:20]


def map_identity_and_contract(job_type_value):
    job_type = sanitize_string(job_type_value)
    if job_type == '1':
        return 'saudi', 'permanent'
    if job_type == '2':
        return 'resident', 'contract'
    if job_type == '4':
        return 'visitor', 'temporary'
    return 'saudi', 'permanent'


def categorize_department(name_ar):
    name = sanitize_string(name_ar)
    if any(keyword in name for keyword in ['طوارئ', 'اسعاف']):
        return 'emergency'
    if any(keyword in name for keyword in ['مختبر', 'أشعة', 'تحاليل', 'صيدلة', 'صيدلية', 'علاج', 'تشخيص']):
        return 'diagnostic'
    if any(keyword in name for keyword in ['صيانة', 'تموين', 'نقل', 'إمداد', 'نظافة', 'ضيافة', 'مغسلة', 'أمن']):
        return 'support'
    if any(keyword in name for keyword in ['تمريض', 'طبيب', 'طبية', 'جراحة', 'عيادة', 'سجلات', 'صحة', 'رعاية']):
        return 'medical'
    return 'administrative'


JOB_TITLE_KEYWORDS = {
    'doctor': ['طبيب', 'طبيبة', 'استشاري', 'استشارية', 'دكتور', 'دكتورة'],
    'nurse': ['ممرض', 'ممرضة', 'تمريض'],
    'technician': ['فني', 'فنية', 'تقني', 'تقنية'],
    'specialist': ['اخصائي', 'أخصائي', 'اخصائية', 'أخصائية', 'اختصاصي', 'اختصاصية'],
    'management': ['مدير', 'مديرة', 'مشرف', 'مشرفة', 'رئيس', 'رئيسة', 'منسق', 'منسقة'],
    'support': ['عامل', 'عاملة', 'سائق', 'حارس', 'مراسل'],
}


def categorize_job_title(title):
    title_value = sanitize_string(title)
    lowered = title_value.lower()
    for category, keywords in JOB_TITLE_KEYWORDS.items():
        for keyword in keywords:
            if keyword in title_value or keyword in lowered:
                return category
    return 'administrative'


def guess_gender(full_name, job_title):
    title = sanitize_string(job_title)
    for keyword in FEMALE_HINTS:
        if keyword in title:
            return 'female'
    for keyword in MALE_HINTS:
        if keyword in title:
            return 'male'
    first_part = sanitize_string(full_name).split()
    if first_part:
        token = first_part[0]
        if token.endswith(('ة', 'ى', 'ا', 'ه')):
            return 'female'
    return 'male'


def split_arabic_name(full_name):
    name = sanitize_string(full_name)
    if not name:
        return ('غير', 'محدد', 'غير', 'محدد')
    parts = [p for p in name.split() if p]
    while len(parts) < 4:
        parts.append('غير')
    first = parts[0]
    second = parts[1] if len(parts) > 1 else 'غير'
    third = parts[2] if len(parts) > 2 else 'غير'
    family = ' '.join(parts[3:]) if len(parts) > 3 else 'محدد'
    if not family:
        family = 'محدد'
    return first, second, third, family


def prepare_email_corrections(rows):
    emails = [sanitize_string(row[7]).lower() for row in rows]
    counts = Counter(email for email in emails if email)
    duplicates = {email for email, count in counts.items() if count > 1}
    return counts, duplicates


def build_department_cache(cursor, dept_code_col):
    cursor.execute(f'SELECT department_id, name_ar, {dept_code_col} FROM Departments')
    cache = {}
    existing_codes = []
    for dept_id, name_ar, code in cursor.fetchall():
        if name_ar:
            cache[sanitize_string(name_ar)] = dept_id
        if code:
            existing_codes.append(sanitize_string(code))
    next_number = 1
    digits = [int(match.group(1)) for code in existing_codes if (match := re.search(r'(\d+)$', code))]
    if digits:
        next_number = max(digits) + 1
    return cache, count(start=next_number)


def build_job_title_cache(cursor, job_titles_table):
    cursor.execute(f'SELECT job_title_id, title_ar, title_code FROM {job_titles_table}')
    cache = {}
    existing_codes = []
    for title_id, title_ar, code in cursor.fetchall():
        if title_ar:
            cache[sanitize_string(title_ar)] = title_id
        if code:
            existing_codes.append(sanitize_string(code))
    next_number = 1
    digits = [int(match.group(1)) for code in existing_codes if (match := re.search(r'(\d+)$', code))]
    if digits:
        next_number = max(digits) + 1
    return cache, count(start=next_number)


def ensure_department(cursor, cache, counter, name_ar, dept_code_col, has_department_type):
    key = sanitize_string(name_ar) or 'قسم غير محدد'
    if key in cache:
        return cache[key]
    code = f"DEPT{next(counter):04d}"
    dept_type = categorize_department(key)
    if has_department_type:
        cursor.execute(
            f'INSERT INTO Departments ({dept_code_col}, name_en, name_ar, description, department_type, is_active) VALUES (%s, %s, %s, %s, %s, 1)',
            (code, f'Department {code}', key, None, dept_type)
        )
    else:
        cursor.execute(
            f'INSERT INTO Departments ({dept_code_col}, name_en, name_ar, description, is_active) VALUES (%s, %s, %s, %s, 1)',
            (code, f'Department {code}', key, dept_type)
        )
    dept_id = cursor.lastrowid
    cache[key] = dept_id
    return dept_id


def ensure_job_title(cursor, cache, counter, title_ar, job_titles_table):
    key = sanitize_string(title_ar) or 'مسمى وظيفي غير محدد'
    if key in cache:
        return cache[key]
    code = f"JT{next(counter):04d}"
    category = categorize_job_title(key)
    cursor.execute(
        f'INSERT INTO {job_titles_table} (title_code, title_en, title_ar, category, is_active) VALUES (%s, %s, %s, %s, 1)',
        (code, f'Job Title {code}', key, category)
    )
    title_id = cursor.lastrowid
    cache[key] = title_id
    return title_id


def chunked(iterable, size):
    current = []
    for item in iterable:
        current.append(item)
        if len(current) == size:
            yield current
            current = []
    if current:
        yield current


def perform_legacy_import(cursor, connection, rows, email_counts, dept_code_col, has_department_type, job_titles_table):
    department_cache, department_counter = build_department_cache(cursor, dept_code_col)
    job_title_cache, job_title_counter = build_job_title_cache(cursor, job_titles_table)

    col_order = [
        'employee_number', 'full_name_en', 'full_name_ar', 'first_name', 'middle_name', 'last_name',
        'department_id', 'position', 'job_title', 'phone_primary', 'phone_secondary', 'email_work', 'hire_date'
    ]

    employees_payload = []
    base_birth = date(1980, 1, 1)
    base_hire = date(2020, 1, 1)

    for index, row in enumerate(rows, start=1):
        full_name = sanitize_string(row[0])
        national_id = sanitize_string(row[1])
        employee_number = sanitize_string(row[2])
        nationality = sanitize_string(row[3])
        department_name = sanitize_string(row[4])
        job_title_name = sanitize_string(row[5])
        phone_number = sanitize_phone(row[6])
        email_value = sanitize_string(row[7]).lower()
        job_type_value = sanitize_string(row[8])

        email_to_use = email_value if email_value and email_value not in PLACEHOLDER_EMAILS and email_counts[email_value] == 1 else None

        first_ar, second_ar, third_ar, family_ar = split_arabic_name(full_name)
        department_id = ensure_department(cursor, department_cache, department_counter, department_name, dept_code_col, has_department_type)
        ensure_job_title(cursor, job_title_cache, job_title_counter, job_title_name, job_titles_table)

        hire_date = base_hire + timedelta(days=index % 365)

        employee_data = {
            'employee_number': employee_number,
            'full_name_en': None,
            'full_name_ar': full_name or f'���� {index}',
            'first_name': None,
            'middle_name': None,
            'last_name': None,
            'department_id': department_id,
            'position': categorize_job_title(job_title_name),
            'job_title': job_title_name,
            'phone_primary': phone_number,
            'phone_secondary': None,
            'email_work': email_to_use,
            'hire_date': hire_date,
        }
        employees_payload.append(tuple(employee_data.get(field) for field in col_order))

    insert_sql = 'INSERT INTO Employees (' + ', '.join(col_order) + ') VALUES (' + ', '.join(['%s'] * len(col_order)) + ')'

    batch_size = 200
    for batch_index, batch in enumerate(chunked(employees_payload, batch_size), start=1):
        cursor.executemany(insert_sql, batch)
        connection.commit()
        print(f'Inserted legacy batch {batch_index} with {len(batch)} employees')

    # Create App_Users for employees with email addresses
    print('Creating App_Users for employees with email addresses...')
    
    # Get all employees with email addresses that were just inserted
    cursor.execute("""
        SELECT employee_id, employee_number, full_name_ar, email_work, job_title, position
        FROM Employees 
        WHERE email_work IS NOT NULL AND email_work != ''
    """)
    
    employees_with_email = cursor.fetchall()
    print(f'Found {len(employees_with_email)} employees with email addresses')
    
    # Create App_Users records
    app_users_data = []
    for emp_id, emp_number, full_name, email, job_title, position in employees_with_email:
        # Determine role based on job title/position
        role = 'employee'  # Default role
        job_lower = (job_title or '').lower()
        position_lower = (position or '').lower()
        
        if any(keyword in job_lower or keyword in position_lower for keyword in ['مدير', 'رئيس', 'مشرف']):
            role = 'manager'
        elif any(keyword in job_lower or keyword in position_lower for keyword in ['موارد بشرية', 'hr']):
            role = 'hr'
        elif any(keyword in job_lower or keyword in position_lower for keyword in ['تقنية', 'it', 'حاسوب']):
            role = 'it'
        elif any(keyword in job_lower or keyword in position_lower for keyword in ['مالية', 'محاسب', 'finance']):
            role = 'finance'
        
        # Default password (employees will need to change)
        default_password = '$2b$12$CqJX9oIoYNewc19hXtODHe4qzM7gc3KJ2YMBX1IzBgE4c3OJbGdvG'  # "Admin@123"
        
        app_users_data.append((
            full_name or f'موظف {emp_number}',  # name
            email,                              # email
            default_password,                   # password_hash
            role,                              # role
            emp_id,                            # employee_id
            1                                  # is_active
        ))
    
    # Insert App_Users in batches
    if app_users_data:
        app_users_sql = """
            INSERT IGNORE INTO App_Users (name, email, password_hash, role, employee_id, is_active)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        
        for batch_index, batch in enumerate(chunked(app_users_data, batch_size), start=1):
            cursor.executemany(app_users_sql, batch)
            connection.commit()
            print(f'Inserted App_Users batch {batch_index} with {len(batch)} users')
    
    print(f'Created {len(app_users_data)} App_Users records for employees with email addresses')


def main():
    excel_path = locate_excel()
    headers, rows = load_rows(excel_path)
    if len(headers) < 8:
        raise ValueError('Unexpected spreadsheet format; at least 8 columns expected')

    email_counts, duplicate_emails = prepare_email_corrections(rows)

    connection = mc.connect(
        host='localhost',
        user='nora',
        password='nora123',
        database='hospital_management'
    )
    connection.autocommit = False
    cursor = connection.cursor()

    # Detect schema differences across deployments
    cursor.execute('SHOW COLUMNS FROM Departments')
    dept_cols = {row[0] for row in cursor.fetchall()}
    dept_code_col = 'department_code' if 'department_code' in dept_cols else 'code'
    has_department_type = 'department_type' in dept_cols

    # Determine job titles table name
    job_titles_table = 'Job_Titles'
    try:
        cursor.execute('SELECT 1 FROM Job_Titles LIMIT 1')
        cursor.fetchall()
    except Exception:
        job_titles_table = 'job_titles'

    # Legacy Employees schema path
    cursor.execute('SHOW COLUMNS FROM Employees')
    emp_cols = {row[0] for row in cursor.fetchall()}
    if 'first_name_ar' not in emp_cols:
        perform_legacy_import(cursor, connection, rows, email_counts, dept_code_col, has_department_type, job_titles_table)
        cursor.close()
        connection.close()
        print(f'Imported {len(rows)} employee records from {excel_path}')
        return

    department_cache, department_counter = build_department_cache(cursor, dept_code_col)
    job_title_cache, job_title_counter = build_job_title_cache(cursor, job_titles_table)

    employees_payload = []
    base_birth = date(1980, 1, 1)
    base_hire = date(2020, 1, 1)

    for index, row in enumerate(rows, start=1):
        full_name = sanitize_string(row[0])
        national_id = sanitize_string(row[1])
        employee_number = sanitize_string(row[2])
        nationality = sanitize_string(row[3]) or 'غير محدد'
        department_name = sanitize_string(row[4])
        job_title_name = sanitize_string(row[5])
        phone_number = sanitize_phone(row[6])
        email_value = sanitize_string(row[7]).lower()
        job_type_value = sanitize_string(row[8])

        email_to_use = email_value if email_value and email_value not in PLACEHOLDER_EMAILS and email_counts[email_value] == 1 else None

        identity_type, contract_type = map_identity_and_contract(job_type_value)
        gender = guess_gender(full_name, job_title_name)
        first_ar, second_ar, third_ar, family_ar = split_arabic_name(full_name)

        department_id = ensure_department(cursor, department_cache, department_counter, department_name, dept_code_col, has_department_type)
        job_title_id = ensure_job_title(cursor, job_title_cache, job_title_counter, job_title_name, job_titles_table)

        birth_date = base_birth + timedelta(days=index % 1000)
        hire_date = base_hire + timedelta(days=index % 365)

        employees_payload.append((
            employee_number,
            first_ar,
            second_ar,
            third_ar,
            family_ar,
            full_name or f'موظف {index}',
            national_id,
            identity_type,
            nationality,
            gender,
            birth_date,
            phone_number,
            email_to_use,
            department_id,
            job_title_id,
            hire_date,
            contract_type,
            'active',
            hire_date
        ))

    insert_sql = (
        'INSERT INTO Employees ('
        'employee_number, first_name_ar, second_name_ar, third_name_ar, family_name_ar, '
        'full_name_ar, national_id, identity_type, nationality, gender, birth_date, '
        'phone_primary, email_work, department_id, job_title_id, hire_date, contract_type, '
        'employment_status, contract_start_date'
        ') VALUES (' + ', '.join(['%s'] * 19) + ')'
    )

    batch_size = 200
    for batch_index, batch in enumerate(chunked(employees_payload, batch_size), start=1):
        cursor.executemany(insert_sql, batch)
        connection.commit()
        print(f'Inserted batch {batch_index} with {len(batch)} employees')

    # Create App_Users for employees with email addresses
    print('Creating App_Users for employees with email addresses...')
    
    # Get all employees with email addresses that were just inserted
    cursor.execute("""
        SELECT employee_id, employee_number, full_name_ar, email_work, job_title, position
        FROM Employees 
        WHERE email_work IS NOT NULL AND email_work != ''
    """)
    
    employees_with_email = cursor.fetchall()
    print(f'Found {len(employees_with_email)} employees with email addresses')
    
    # Create App_Users records
    app_users_data = []
    for emp_id, emp_number, full_name, email, job_title, position in employees_with_email:
        # Determine role based on job title/position
        role = 'employee'  # Default role
        job_lower = (job_title or '').lower()
        position_lower = (position or '').lower()
        
        if any(keyword in job_lower or keyword in position_lower for keyword in ['مدير', 'رئيس', 'مشرف']):
            role = 'manager'
        elif any(keyword in job_lower or keyword in position_lower for keyword in ['موارد بشرية', 'hr']):
            role = 'hr'
        elif any(keyword in job_lower or keyword in position_lower for keyword in ['تقنية', 'it', 'حاسوب']):
            role = 'it'
        elif any(keyword in job_lower or keyword in position_lower for keyword in ['مالية', 'محاسب', 'finance']):
            role = 'finance'
        
        # Default password (employees will need to change)
        default_password = '$2b$12$CqJX9oIoYNewc19hXtODHe4qzM7gc3KJ2YMBX1IzBgE4c3OJbGdvG'  # "Admin@123"
        
        app_users_data.append((
            full_name or f'موظف {emp_number}',  # name
            email,                              # email
            default_password,                   # password_hash
            role,                              # role
            emp_id,                            # employee_id
            1                                  # is_active
        ))
    
    # Insert App_Users in batches
    if app_users_data:
        app_users_sql = """
            INSERT IGNORE INTO App_Users (name, email, password_hash, role, employee_id, is_active)
            VALUES (%s, %s, %s, %s, %s, %s)
        """
        
        for batch_index, batch in enumerate(chunked(app_users_data, batch_size), start=1):
            cursor.executemany(app_users_sql, batch)
            connection.commit()
            print(f'Inserted App_Users batch {batch_index} with {len(batch)} users')
    
    print(f'Created {len(app_users_data)} App_Users records for employees with email addresses')

    cursor.close()
    connection.close()
    print(f'Imported {len(employees_payload)} employee records and created {len(app_users_data)} user accounts from {excel_path}')


if __name__ == '__main__':
    main()
